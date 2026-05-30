"""Экспорт результатов в Excel"""

import os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from model.parameters import TaskParameters
from model.results import OptimizationResults


def export_to_excel(results: OptimizationResults, params: TaskParameters, filepath: str):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлен. pip install openpyxl")

    wb = openpyxl.Workbook()

    # --- Лист 1: Параметры ---
    ws1 = wb.active
    ws1.title = "Параметры"
    _write_params(ws1, params)

    # --- Лист 2: Составы пакетов ---
    ws2 = wb.create_sheet("Составы пакетов")
    _write_batches(ws2, results)

    # --- Лист 3: Расписание ---
    ws3 = wb.create_sheet("Расписание")
    _write_schedule(ws3, results, params)

    # --- Лист 4: ТО ---
    if results.maintenance:
        ws4 = wb.create_sheet("Техническое обслуживание")
        _write_maintenance(ws4, results)

    # --- Лист 5: Сводка ---
    ws5 = wb.create_sheet("Сводка")
    _write_summary(ws5, results, params)

    wb.save(filepath)
    return filepath


def _header_style():
    return Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="2F5597")


def _apply_header(cell, text):
    font, fill = _header_style()
    cell.value = text
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _write_params(ws, params: TaskParameters):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    bold = Font(bold=True)
    rows = [
        ("ПАРАМЕТРЫ ЗАДАЧИ", ""),
        ("", ""),
        ("Количество типов заданий (I)", params.I),
        ("Количество приборов (L)", params.L),
        ("Количество позиций (J)", params.J),
        ("", ""),
        ("Количество заданий по типам", ""),
    ]
    for r in rows:
        ws.append(r)
    ws['A1'].font = Font(bold=True, size=14)

    for i in range(params.I):
        ws.append([f"  n[{i+1}]", params.n[i]])

    ws.append(["", ""])
    ws.append(["Времена обработки t[прибор][тип]", ""])

    header = ["Прибор \\ Тип"] + [f"Тип {i+1}" for i in range(params.I)]
    ws.append(header)
    for l in range(params.L):
        row = [f"Прибор {l+1}"] + params.t[l]
        ws.append(row)

    ws.append(["", ""])
    ws.append(["Директивные сроки", ""])
    for i in range(params.I):
        ws.append([f"  d[{i+1}]", params.d[i]])

    if params.use_maintenance:
        ws.append(["", ""])
        ws.append(["Параметры ТО", ""])
        ws.append(["Прибор", "Период TM", "Длительность tm"])
        for l in range(params.L):
            ws.append([f"Прибор {l+1}", params.TM[l], params.tm_maint[l]])


def _write_batches(ws, results: OptimizationResults):
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

    headers = ["Позиция j", "Тип заданий i", "Количество заданий mj"]
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(1, col), h)

    for row, b in enumerate(sorted(results.batches, key=lambda x: x.position), 2):
        ws.cell(row, 1, b.position)
        ws.cell(row, 2, b.task_type)
        ws.cell(row, 3, b.count)


def _write_schedule(ws, results: OptimizationResults, params: TaskParameters):
    headers = ["Прибор l", "Позиция j", "Тип заданий", "Начало q_lj",
               "Время обработки", "Окончание"]
    ws.column_dimensions['A'].width = 12
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(1, col), h)

    sorted_sched = sorted(results.schedule, key=lambda e: (e.device, e.position))
    for row, e in enumerate(sorted_sched, 2):
        ws.cell(row, 1, e.device)
        ws.cell(row, 2, e.position)
        ws.cell(row, 3, e.task_type)
        ws.cell(row, 4, round(e.start_time, 4))
        ws.cell(row, 5, round(e.process_time, 4))
        ws.cell(row, 6, round(e.end_time, 4))


def _write_maintenance(ws, results: OptimizationResults):
    headers = ["Прибор l", "Перед позицией", "Начало ТО", "Длительность", "Окончание ТО"]
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(1, col), h)
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row, m in enumerate(sorted(results.maintenance, key=lambda x: (x.device, x.start_time)), 2):
        ws.cell(row, 1, m.device)
        ws.cell(row, 2, m.before_position)
        ws.cell(row, 3, round(m.start_time, 4))
        ws.cell(row, 4, round(m.duration, 4))
        ws.cell(row, 5, round(m.end_time, 4))


def _write_summary(ws, results: OptimizationResults, params: TaskParameters):
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25

    ws['A1'] = "СВОДКА РЕЗУЛЬТАТОВ ОПТИМИЗАЦИИ"
    ws['A1'].font = Font(bold=True, size=14)

    data = [
        ("Статус решения", results.status),
        ("Критерий оптимизации", results.criterion),
        ("Значение критерия (опт.)", round(results.objective_value, 4)
         if results.objective_value else "—"),
        ("Значение критерия (фикс. пакеты)", round(results.fixed_objective, 4)
         if results.fixed_objective else "—"),
        ("Улучшение, %", round(results.improvement_percent, 2)
         if results.improvement_percent else "—"),
        ("Время решения, с", round(results.solve_time, 2)),
        ("Решатель", results.solver_name),
        ("Учёт ТО", "Да" if params.use_maintenance else "Нет"),
        ("Количество пакетов", len(results.batches)),
        ("Количество сеансов ТО", len(results.maintenance)),
    ]

    for i, (label, value) in enumerate(data, 3):
        ws.cell(i, 1, label).font = Font(bold=True)
        ws.cell(i, 2, value)

    if results.delays:
        ws.cell(len(data)+4, 1, "Запаздывания по типам заданий").font = Font(bold=True)
        for i, (k, v) in enumerate(results.delays.items(), len(data)+5):
            ws.cell(i, 1, f"Тип {k}")
            ws.cell(i, 2, round(v, 4))
